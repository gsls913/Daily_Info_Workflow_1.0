"""
AI内容生成模块 - AlphaPai会议纪要
读取Excel标签选项，调用AI判断行业和公司标签
支持并行处理多个文件
"""
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    openpyxl = None

from .ai_client import AIClient, create_ai_client, get_parallel_workers


# 行业标签Prompt模板
INDUSTRY_PROMPT_TEMPLATE = """你是一位专业的投资研究员，擅长分析会议纪要并提取关键行业信息。

请仔细阅读以下会议纪要要点，判断该纪要主要涉及哪些行业或领域。

【可选行业/领域列表】
{options}

【会议纪要要点】
{text}

【判断要求】
1. **选择标准**：只有会议纪要中确实有一定篇幅详细讨论某个行业/领域时，才能选择该标签。不能仅因为提到某个行业就选择。

2. **标签类型说明**：
   - 狭义行业标签：如酒店、餐饮、零售、汽车、医药等具体行业
   - 广义类型标签：如宏观、策略、国际政治、消费行业整体、周期股等投资类型属性
   - 两者可以同时选择，根据内容判断，但一定不要在并未涉及某个标签时选择该标签

3. **数量控制**：
   - 选择0-3个最相关的标签，不要超过3个
   - 如果纪要没有明确涉及上述列表中的任何行业/领域，请返回"无"

4. **输出格式**：
   - 多个标签用顿号"、"分隔，如：酒店、餐饮、宏观
   - 不要输出任何解释性文字，只输出标签名称
   - 如果没有匹配的标签，只输出"无"，请勿输出可选标签列表以外的标签

请直接输出判断结果："""


# 公司标签Prompt模板
COMPANY_PROMPT_TEMPLATE = """你是一位专业的投资研究员，擅长分析会议纪要并提取关键公司信息。

请仔细阅读以下会议纪要要点，判断该纪要主要涉及哪些公司。

【可选公司列表】
{options}

【会议纪要要点】
{text}

【判断要求】
1. **选择标准**：
   - 只有会议纪要中确实有一定篇幅详细讨论某家公司时，才能选择该公司标签
   - 不能仅因为提到某家公司就选择，必须是重点讨论的对象
   - 如果内容大部分围绕某一家公司展开（即使少数地方提及其他公司），请选择那个主要公司
   - 如果内容明确对多家公司都有重点描写，可以选择多家，但一定不要在并未涉及某个公司时选择该公司

2. **数量控制**：
   - 通常选择0-3个最相关的公司
   - 如果纪要确实详细讨论了较多公司，可以适当增加（最多不超过5个）
   - 如果纪要没有明确讨论上述列表中的任何公司，请返回"无"

3. **输出格式**：
   - 多个公司用顿号"、"分隔，如：中国中免、名创优品、锦江酒店
   - 不要输出任何解释性文字，只输出公司名称
   - 如果没有匹配的公司，只输出"无"，请勿输出可选公司列表以外的公司

请直接输出判断结果："""


def load_tag_options(excel_path: Optional[Path] = None) -> Dict[str, List[str]]:
    """
    从Excel文件加载标签选项
    
    Args:
        excel_path: Excel文件路径，默认为 data/config/set_config.xlsx 的 memo_tag_options sheet
    
    Returns:
        Dict[str, List[str]]: {"行业": [...], "公司": [...]}
    """
    if excel_path is None:
        excel_path = Path(__file__).resolve().parent.parent.parent / "data" / "config" / "set_config.xlsx"
    
    if not excel_path.exists():
        raise FileNotFoundError(f"标签选项文件不存在: {excel_path}")
    
    if openpyxl is None:
        raise ImportError("请先安装openpyxl: pip install openpyxl")
    
    result = {"行业": [], "公司": []}
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        
        sheet = None
        if "memo_tag_options" in wb.sheetnames:
            sheet = wb["memo_tag_options"]
        else:
            raise RuntimeError(f"Excel文件中未找到 'memo_tag_options' sheet")
        
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        industry_col = None
        company_col = None
        for idx, header in enumerate(headers, 1):
            if header == "行业":
                industry_col = idx
            elif header == "公司":
                company_col = idx
        
        max_row = sheet.max_row
        
        if industry_col:
            for row in range(2, max_row + 1):
                cell = sheet.cell(row=row, column=industry_col)
                if cell.value:
                    value = str(cell.value).strip()
                    if value and value not in result["行业"]:
                        result["行业"].append(value)
        
        if company_col:
            for row in range(2, max_row + 1):
                cell = sheet.cell(row=row, column=company_col)
                if cell.value:
                    value = str(cell.value).strip()
                    if value and value not in result["公司"]:
                        result["公司"].append(value)
        
        wb.close()
        
    except Exception as e:
        raise RuntimeError(f"读取Excel文件失败: {e}")
    
    return result


def extract_ai_summary(md_content: str) -> str:
    """
    从Markdown内容中提取AI要点部分
    
    Args:
        md_content: Markdown文件内容
    
    Returns:
        AI要点部分的文本，如果找不到返回空字符串
    """
    # 查找"# AI 要点"或"## AI 要点"等标题
    patterns = [
        r'# AI 要点\s*\n(.*?)(?=\n# |\Z)',  # 一级标题
        r'## AI 要点\s*\n(.*?)(?=\n## |\n# |\Z)',  # 二级标题
        r'# AI要点\s*\n(.*?)(?=\n# |\Z)',  # 无空格
        r'## AI要点\s*\n(.*?)(?=\n## |\n# |\Z)',  # 无空格
    ]
    
    for pattern in patterns:
        match = re.search(pattern, md_content, re.DOTALL | re.IGNORECASE)
        if match:
            return match.group(1).strip()
    
    # 如果没找到，尝试找"AI 要点"或"AI要点"字样
    alt_patterns = [
        r'AI 要点\s*\n(.*?)(?=\n# |\n---|\Z)',
        r'AI要点\s*\n(.*?)(?=\n# |\n---|\Z)',
    ]
    
    for pattern in alt_patterns:
        match = re.search(pattern, md_content, re.DOTALL | re.IGNORECASE)
        if match:
            return match.group(1).strip()
    
    # 如果找不到AI要点，返回空字符串（不再使用fallback）
    return ""


def generate_tags_for_meeting(
    md_file_path: Path,
    ai_client: AIClient,
    tag_options: Dict[str, List[str]],
    log_func = None
) -> Tuple[List[str], List[str], float, float]:
    """
    为单个会议纪要文件生成标签
    
    Args:
        md_file_path: Markdown文件路径
        ai_client: AI客户端实例
        tag_options: 标签选项字典
        log_func: 日志函数
    
    Returns:
        Tuple[行业标签列表, 公司标签列表, 行业标签用时秒数, 公司标签用时秒数]
    """
    if log_func is None:
        log_func = lambda msg, level: print(f"[{level}] {msg}")
    
    # 读取文件内容
    try:
        content = md_file_path.read_text(encoding="utf-8")
    except Exception as e:
        log_func(f"读取文件失败 {md_file_path}: {e}", "ERROR")
        return [], [], 0.0, 0.0
    
    # 提取AI要点部分
    ai_summary = extract_ai_summary(content)
    if not ai_summary:
        log_func(f"未找到AI要点部分: {md_file_path.name}", "WARN")
        return [], [], 0.0, 0.0
    
    # 限制长度，避免token过多
    ai_summary = ai_summary[:2500]
    
    industry_tags = []
    company_tags = []
    industry_elapsed = 0.0
    company_elapsed = 0.0
    
    # 判断行业标签
    if tag_options.get("行业"):
        try:
            start_time = time.time()
            tags, metadata = ai_client.call_for_tags(
                text=ai_summary,
                property_name="行业",
                valid_options=tag_options["行业"],
                prompt_template=INDUSTRY_PROMPT_TEMPLATE,
                max_attempts_per_model=3,
                max_models=3
            )
            industry_elapsed = time.time() - start_time
            
            if tags is not None:
                industry_tags = tags
            else:
                log_func(f"行业标签判断失败: {metadata.get('error')}", "WARN")
        except Exception as e:
            log_func(f"行业标签判断异常: {e}", "ERROR")
    
    # 判断公司标签
    if tag_options.get("公司"):
        try:
            start_time = time.time()
            tags, metadata = ai_client.call_for_tags(
                text=ai_summary,
                property_name="公司",
                valid_options=tag_options["公司"],
                prompt_template=COMPANY_PROMPT_TEMPLATE,
                max_attempts_per_model=3,
                max_models=3
            )
            company_elapsed = time.time() - start_time
            
            if tags is not None:
                company_tags = tags
            else:
                log_func(f"公司标签判断失败: {metadata.get('error')}", "WARN")
        except Exception as e:
            log_func(f"公司标签判断异常: {e}", "ERROR")
    
    return industry_tags, company_tags, industry_elapsed, company_elapsed


def update_md_tags(md_file_path: Path, industry_tags: List[str], company_tags: List[str]) -> bool:
    """
    更新Markdown文件中的标签（限定在基本信息部分）
    
    Args:
        md_file_path: Markdown文件路径
        industry_tags: 行业标签列表
        company_tags: 公司标签列表
    
    Returns:
        是否成功更新
    """
    try:
        content = md_file_path.read_text(encoding="utf-8")
        
        # 构建标签字符串
        industry_str = " ".join([f"#{tag}" for tag in industry_tags]) if industry_tags else ""
        company_str = " ".join([f"#{tag}" for tag in company_tags]) if company_tags else ""
        
        # 找到基本信息部分的边界
        basic_info_start = content.find("# 基本信息")
        if basic_info_start == -1:
            basic_info_start = 0
        
        # 找到基本信息部分的结束位置（下一个一级标题或分隔线）
        basic_info_end = len(content)
        next_h1 = content.find("\n# ", basic_info_start + 1)
        if next_h1 != -1:
            basic_info_end = next_h1
        next_separator = content.find("\n---", basic_info_start + 1)
        if next_separator != -1 and next_separator < basic_info_end:
            basic_info_end = next_separator
        
        basic_info_section = content[basic_info_start:basic_info_end]
        
        # 在基本信息部分内查找并替换行业标签行
        industry_pattern = r'(- \*\*行业\*\*:).*'
        if re.search(industry_pattern, basic_info_section):
            new_basic_info = re.sub(industry_pattern, f'\\1 {industry_str}', basic_info_section)
        else:
            # 如果没有找到，在基本信息部分末尾添加
            new_basic_info = basic_info_section.rstrip() + f"\n- **行业**: {industry_str}"
        
        # 在基本信息部分内查找并替换公司标签行
        company_pattern = r'(- \*\*公司\*\*:).*'
        if re.search(company_pattern, new_basic_info):
            new_basic_info = re.sub(company_pattern, f'\\1 {company_str}', new_basic_info)
        else:
            # 如果没有找到，在行业标签后添加
            new_basic_info = new_basic_info.rstrip() + f"\n- **公司**: {company_str}"
        
        # 重新组合内容
        new_content = content[:basic_info_start] + new_basic_info + content[basic_info_end:]
        
        # 写回文件
        md_file_path.write_text(new_content, encoding="utf-8")
        return True
        
    except Exception as e:
        print(f"更新文件标签失败 {md_file_path}: {e}")
        return False


# ============================================================
# AI 评价功能
# ============================================================
ANALYSIS_PROMPT_TEMPLATE = """你是一位资深的行业研究员，专注于市场、股票和行业研究。请对以下会议纪要进行深入分析和评价。

【原文内容】
{text}

请从以下几个维度进行分析（总字数控制在600-900字）：

## 一、深度解读与思辨

请从以下角度中选择适合本文的3-5个角度进行深入分析（不必全部使用，选择最相关的即可）：

1. **核心信息**：会议的核心议题和关键结论是什么？
2. **管理层观点**：管理层对行业和公司的判断是否合理？有什么隐含信息？
3. **行业趋势**：透露了哪些行业发展趋势和竞争格局变化？
4. **风险与机会**：有哪些潜在风险和投资机会？
5. **历史/类比参考**：是否有类似的历史案例可以类比？

## 二、值得关注的信息点

请列出3-5个特别值得关注的信息点或数据，并简要说明原因。
（请使用列表形式，如：- 信息点1：说明原因）

## 三、下一步跟踪验证

对于原文涉及的投资机会或风险，请列出需要进一步跟踪验证的关键指标或潜在事件：
- 哪些关键数据/指标需要持续观察？
- 哪些潜在事件可能改变判断？
- 需要进一步研究什么问题？

【输出要求】
- 语言简洁专业，避免空泛表述
- 观点要有逻辑支撑，体现思辨性
- 选择最适合本文的分析角度，不必强行使用所有角度
- 如信息量不足，可适当说明局限性
- **禁止使用表格**，请使用列表或段落形式呈现内容
- 可以使用Markdown格式（如 ## 标题、**粗体**、- 列表 等）"""


def upgrade_headings(text: str) -> str:
    """
    提升标题层级：如果最高层级是一级标题，则所有标题都提升一级
    
    Args:
        text: 原始文本
    
    Returns:
        处理后的文本
    """
    lines = text.split('\n')
    
    # 找出所有标题层级
    heading_levels = []
    for line in lines:
        match = re.match(r'^(#{1,6})\s+', line)
        if match:
            heading_levels.append(len(match.group(1)))
    
    if not heading_levels:
        return text
    
    # 找出最高层级（数字最小）
    min_level = min(heading_levels)
    
    # 如果最高层级是一级标题，则所有标题都提升一级
    if min_level == 1:
        result_lines = []
        for line in lines:
            match = re.match(r'^(#{1,5})(\s+.*)$', line)
            if match:
                current_level = len(match.group(1))
                rest = match.group(2)
                # 提升一级（加一个#）
                new_line = '#' + match.group(1) + rest
                result_lines.append(new_line)
            else:
                result_lines.append(line)
        return '\n'.join(result_lines)
    
    return text


def extract_thinking_content(text: str) -> str:
    """
    提取最终答案，去除思考过程（MiniMax 模型可能返回思考标签）
    
    Args:
        text: 原始响应文本
    
    Returns:
        处理后的文本
    """
    text = text.strip()
    
    thinking_start_tag = ""
    thinking_end_tag = ""
    
    if thinking_start_tag in text or thinking_end_tag in text:
        thinking_end = text.find(thinking_end_tag)
        if thinking_end != -1:
            text = text[thinking_end + len(thinking_end_tag):].strip()
        else:
            thinking_start = text.find(thinking_start_tag)
            if thinking_start != -1:
                remaining = text[thinking_start + len(thinking_start_tag):]
                lines = remaining.split('\n')
                for line in lines:
                    line = line.strip()
                    if line and not line.startswith('<') and not line.startswith('思考'):
                        text = line
                        break
    
    return text


def generate_ai_analysis(
    md_file_path: Path,
    ai_client: "AIClient",
    log_func = None
) -> Tuple[str, float]:
    """
    为单个会议纪要文件生成AI评价
    
    Args:
        md_file_path: Markdown文件路径
        ai_client: AI客户端实例
        log_func: 日志函数
    
    Returns:
        Tuple[AI评价内容, 用时秒数]
    """
    start_time = time.time()
    
    if log_func is None:
        log_func = lambda msg, level: print(f"[{level}] {msg}")
    
    # 读取文件内容
    try:
        content = md_file_path.read_text(encoding="utf-8")
    except Exception as e:
        log_func(f"读取文件失败 {md_file_path}: {e}", "ERROR")
        return "", 0.0
    
    # 提取AI要点部分
    ai_summary = extract_ai_summary(content)
    if not ai_summary:
        log_func(f"未找到AI要点部分: {md_file_path.name}", "WARN")
        return "", 0.0
    
    # 限制长度
    ai_summary = ai_summary[:4000]
    
    # 构建prompt
    prompt = ANALYSIS_PROMPT_TEMPLATE.format(text=ai_summary)
    
    # 调用AI（使用long_thinking模式）
    try:
        response, metadata = ai_client.call_for_long_thinking(
            prompt=prompt,
            temperature=0.7,
            max_tokens=2000,
            max_attempts_per_model=3,
            max_models=3
        )
        
        # 去除思考过程（MiniMax 可能返回思考标签）
        response = extract_thinking_content(response)
        
        # 处理标题层级
        processed_response = upgrade_headings(response)
        
        elapsed = time.time() - start_time
        
        return processed_response, elapsed
        
    except Exception as e:
        log_func(f"AI评价生成失败: {e}", "ERROR")
        return "", 0.0


def insert_ai_analysis_to_md(md_file_path: Path, ai_analysis: str) -> bool:
    """
    将AI评价插入到Markdown文件中
    
    插入位置优先级：
    1. 如果有"# 会议全文"，插入到它之前
    2. 如果没有，插入到末尾标识（*下载时间:*）之前
    
    Args:
        md_file_path: Markdown文件路径
        ai_analysis: AI评价内容
    
    Returns:
        是否成功更新
    """
    if not ai_analysis:
        return False
    
    try:
        content = md_file_path.read_text(encoding="utf-8")
        
        # 构建AI评价部分
        ai_section = f"\n---\n\n# AI 评价\n\n{ai_analysis}\n"
        
        # 优先查找"# 会议全文"的位置
        meeting_full_text_pattern = r'(\n---\s*\n)?(# 会议全文)'
        match = re.search(meeting_full_text_pattern, content)
        
        if match:
            # 在"# 会议全文"之前插入AI评价
            insert_position = match.start()
            new_content = content[:insert_position] + ai_section + content[insert_position:]
        else:
            # 查找末尾标识的位置（*下载时间:）
            download_time_pattern = r'\n---\s*\n\*下载时间:'
            footer_match = re.search(download_time_pattern, content)
            
            if footer_match:
                # 在末尾标识之前插入AI评价
                insert_position = footer_match.start()
                new_content = content[:insert_position] + ai_section + content[insert_position:]
            else:
                # 如果都没有，在文件末尾添加
                new_content = content + ai_section
        
        # 写回文件
        md_file_path.write_text(new_content, encoding="utf-8")
        return True
            
    except Exception as e:
        print(f"插入AI评价失败 {md_file_path}: {e}")
        return False


# 全局计数器和锁（用于并行处理时的进度显示）
_progress_lock = threading.Lock()
_progress_counter = 0


def _process_single_file(
    md_file: Path,
    ai_client: AIClient,
    tag_options: Dict[str, List[str]],
    total_files: int,
    log_func,
    delay_seconds: float = 0
) -> Tuple[str, List[str], List[str], bool, float, float, float]:
    """
    处理单个文件的内部函数（用于并行调用）
    
    Args:
        md_file: 文件路径
        ai_client: AI客户端（线程安全）
        tag_options: 标签选项
        total_files: 总文件数
        log_func: 日志函数
        delay_seconds: 启动延迟秒数
    
    Returns:
        (文件路径字符串, 行业标签, 公司标签, 是否成功添加AI评价, 行业标签用时, 公司标签用时, 评价用时)
    """
    global _progress_counter
    
    # 启动延迟（避免同时发出大量请求）
    if delay_seconds > 0:
        time.sleep(delay_seconds)
    
    # 生成标签
    industry_tags, company_tags, industry_elapsed, company_elapsed = generate_tags_for_meeting(
        md_file, ai_client, tag_options, log_func
    )
    
    # 更新标签
    update_md_tags(md_file, industry_tags, company_tags)
    
    # 生成AI评价
    ai_analysis, analysis_elapsed = generate_ai_analysis(md_file, ai_client, log_func)
    
    # 插入AI评价
    analysis_success = False
    if ai_analysis:
        analysis_success = insert_ai_analysis_to_md(md_file, ai_analysis)
    
    # 更新进度
    with _progress_lock:
        _progress_counter += 1
        ind_str = f"行业{industry_elapsed:.1f}s" if industry_elapsed > 0 else "行业跳过"
        comp_str = f"公司{company_elapsed:.1f}s" if company_elapsed > 0 else "公司跳过"
        analysis_str = f"评价{analysis_elapsed:.1f}s" if analysis_success else "评价跳过"
        log_func(f"[{_progress_counter}/{total_files}] {md_file.name} ({ind_str}, {comp_str}, {analysis_str})", "INFO")
    
    return str(md_file), industry_tags, company_tags, analysis_success, industry_elapsed, company_elapsed, analysis_elapsed


def generate_tags_and_analysis_for_batch_parallel(
    md_files: List[Path],
    log_func = None,
    max_workers: Optional[int] = None,
    return_timing: bool = False
) -> Tuple[Dict[str, Tuple[List[str], List[str], bool]], float, float, float]:
    """
    并行批量为会议纪要文件生成标签和AI评价
    
    使用线程池并行处理多个文件，提高处理效率。
    AI客户端已实现线程安全，支持多线程并发调用。
    
    启动时每个任务依次延迟1秒发出，避免API拥挤。
    
    Args:
        md_files: Markdown文件路径列表
        log_func: 日志函数
        max_workers: 最大并行工作线程数，默认从配置文件读取
        return_timing: 是否返回计时信息
    
    Returns:
        如果 return_timing=True:
            Tuple[结果字典, 行业标签总用时, 公司标签总用时, AI评价总用时]
        否则:
            Dict[文件路径, (行业标签列表, 公司标签列表, 是否成功添加AI评价)]
    """
    global _progress_counter
    _progress_counter = 0
    
    if log_func is None:
        log_func = lambda msg, level: print(f"[{level}] {msg}")
    
    # 获取并行工作线程数
    if max_workers is None:
        max_workers = get_parallel_workers()
    
    log_func(f"并行处理模式: {max_workers} 个工作线程", "INFO")
    
    # 加载标签选项
    try:
        tag_options = load_tag_options()
        log_func(f"已加载标签选项: 行业{len(tag_options['行业'])}个, 公司{len(tag_options['公司'])}个", "INFO")
    except Exception as e:
        log_func(f"加载标签选项失败: {e}", "ERROR")
        return {}
    
    # 创建AI客户端（线程安全）
    try:
        ai_client = create_ai_client(log_func=log_func)
        log_func(f"AI客户端初始化成功 (提供商: {ai_client.get_provider()})", "INFO")
    except Exception as e:
        log_func(f"AI客户端初始化失败: {e}", "ERROR")
        return {}
    
    results = {}
    total_files = len(md_files)
    
    # 统计总用时
    total_industry_time = 0.0
    total_company_time = 0.0
    total_analysis_time = 0.0
    
    log_func(f"开始并行处理 {total_files} 个文件...", "INFO")
    
    # 使用线程池并行处理
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有任务，每个任务依次延迟1秒
        future_to_file = {}
        for idx, md_file in enumerate(md_files):
            delay = idx * 1.0  # 每个任务延迟递增1秒
            future = executor.submit(
                _process_single_file,
                md_file,
                ai_client,
                tag_options,
                total_files,
                log_func,
                delay
            )
            future_to_file[future] = md_file
        
        # 动态获取完成的任务结果
        for future in as_completed(future_to_file):
            md_file = future_to_file[future]
            try:
                file_path, industry_tags, company_tags, analysis_success, ind_time, comp_time, analysis_time = future.result()
                results[file_path] = (industry_tags, company_tags, analysis_success)
                total_industry_time += ind_time
                total_company_time += comp_time
                total_analysis_time += analysis_time
            except Exception as e:
                log_func(f"处理文件失败 {md_file.name}: {e}", "ERROR")
                results[str(md_file)] = ([], [], False)
    
    # 输出统计信息
    log_func(f"处理完成，共 {len(results)} 个文件", "INFO")
    
    if return_timing:
        return results, total_industry_time, total_company_time, total_analysis_time
    return results


if __name__ == "__main__":
    # 测试代码
    print("测试AI内容生成模块...")
    
    # 测试加载标签选项
    try:
        options = load_tag_options()
        print(f"行业标签数量: {len(options['行业'])}")
        print(f"公司标签数量: {len(options['公司'])}")
        print(f"前5个行业标签: {options['行业'][:5]}")
        print(f"前5个公司标签: {options['公司'][:5]}")
    except Exception as e:
        print(f"加载标签选项失败: {e}")
