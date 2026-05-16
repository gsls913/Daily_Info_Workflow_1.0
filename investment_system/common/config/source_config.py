from __future__ import annotations

import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable
from urllib.parse import urlparse

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - exercised by runtime dependency checks
    Workbook = None
    load_workbook = None

from investment_system.common.config.config_loader import get as cfg
from investment_system.common.utils.paths import SET_CONFIG_FILE


WECHAT_SHEET = "wechat_account"
PODCAST_SHEET = "podcast_account"
TAG_SHEET = "memo_tag_options"

WECHAT_HEADERS = ("公众号名称", "简称", "分类", "单篇/聚合")
PODCAST_HEADERS = ("播客名称", "url", "简称")
TAG_HEADERS = ("公司", "行业")


class SourceConfigError(RuntimeError):
    """Raised when set_config.xlsx is missing, malformed, or invalid."""


@dataclass(frozen=True)
class WechatAccount:
    name: str
    short_name: str
    category: str
    mode: str = "单篇"


@dataclass(frozen=True)
class PodcastAccount:
    name: str
    url: str
    short_name: str


def _ensure_openpyxl() -> None:
    if load_workbook is None or Workbook is None:
        raise SourceConfigError("openpyxl未安装，无法读写 data/config/set_config.xlsx，请运行: pip install openpyxl")


def _config_path(path: str | Path | None = None) -> Path:
    return Path(path or SET_CONFIG_FILE)


def _normalize(value) -> str:
    return str(value).strip() if value is not None else ""


def _dedupe_keep_order(values: Iterable[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        item = _normalize(value)
        if item and item not in seen:
            result.append(item)
            seen.add(item)
    return result


def _load_wb(path: str | Path | None = None):
    _ensure_openpyxl()
    file_path = _config_path(path)
    if not file_path.exists():
        raise SourceConfigError(f"配置文件不存在: {file_path}")
    return load_workbook(file_path)


def _save_wb(wb, path: str | Path | None = None, *, backup: bool = True) -> Path | None:
    file_path = _config_path(path)
    file_path.parent.mkdir(parents=True, exist_ok=True)
    backup_path = None
    if backup and file_path.exists():
        stamp = datetime.now().strftime("%Y%m%d%H%M%S")
        backup_path = file_path.with_name(f"{file_path.name}.{stamp}.bak")
        shutil.copy2(file_path, backup_path)
    wb.save(file_path)
    return backup_path


def _header_map(ws, required: Iterable[str]) -> dict[str, int]:
    headers = {_normalize(cell.value): idx for idx, cell in enumerate(ws[1], 1) if _normalize(cell.value)}
    missing = [item for item in required if item not in headers]
    if missing:
        raise SourceConfigError(f"{ws.title} sheet 缺少必要表头: {', '.join(missing)}")
    return headers


def _ensure_sheet(wb, sheet_name: str, headers: Iterable[str]):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row == 0:
            ws.append(list(headers))
        _header_map(ws, headers)
        return ws
    ws = wb.create_sheet(sheet_name)
    ws.append(list(headers))
    return ws


def _valid_wechat_categories() -> set[str]:
    mapping = cfg("wechat.category_mapping", {}) or {}
    return set(mapping.keys()) or {"投资", "宏观", "商业", "科技", "工作", "其他"}


def load_wechat_accounts(path: str | Path | None = None) -> list[dict[str, str]]:
    wb = _load_wb(path)
    try:
        if WECHAT_SHEET not in wb.sheetnames:
            raise SourceConfigError(f"Excel文件中缺少 '{WECHAT_SHEET}' sheet")
        ws = wb[WECHAT_SHEET]
        headers = _header_map(ws, WECHAT_HEADERS[:3])
        categories = _valid_wechat_categories()
        accounts: list[dict[str, str]] = []
        seen: set[str] = set()
        for row_idx in range(2, ws.max_row + 1):
            name = _normalize(ws.cell(row_idx, headers["公众号名称"]).value)
            if not name:
                continue
            short_name = _normalize(ws.cell(row_idx, headers["简称"]).value)
            category = _normalize(ws.cell(row_idx, headers["分类"]).value) or "其他"
            mode_col = headers.get("单篇/聚合")
            mode = _normalize(ws.cell(row_idx, mode_col).value) if mode_col else "单篇"
            if not short_name:
                raise SourceConfigError(f"{WECHAT_SHEET} 第 {row_idx} 行公众号「{name}」缺少简称")
            if category not in categories:
                raise SourceConfigError(f"{WECHAT_SHEET} 第 {row_idx} 行公众号「{name}」分类无效: {category}")
            if name in seen:
                raise SourceConfigError(f"{WECHAT_SHEET} 存在重复公众号名称: {name}")
            seen.add(name)
            accounts.append({"name": name, "short_name": short_name, "category": category, "mode": mode or "单篇"})
        return accounts
    finally:
        wb.close()


def add_wechat_account(
    name: str,
    short_name: str,
    category: str,
    mode: str = "单篇",
    path: str | Path | None = None,
) -> None:
    name = _normalize(name)
    short_name = _normalize(short_name)
    category = _normalize(category)
    mode = _normalize(mode) or "单篇"
    if not name or not short_name:
        raise SourceConfigError("公众号名称和简称不能为空")
    if category not in _valid_wechat_categories():
        raise SourceConfigError(f"公众号分类无效: {category}")
    wb = _load_wb(path)
    try:
        ws = _ensure_sheet(wb, WECHAT_SHEET, WECHAT_HEADERS)
        headers = _header_map(ws, WECHAT_HEADERS)
        existing = {_normalize(ws.cell(row, headers["公众号名称"]).value) for row in range(2, ws.max_row + 1)}
        if name in existing:
            raise SourceConfigError(f"公众号已存在: {name}")
        ws.append([name, short_name, category, mode])
        _save_wb(wb, path)
    finally:
        wb.close()


def remove_wechat_account(name: str, path: str | Path | None = None) -> int:
    name = _normalize(name)
    if not name:
        raise SourceConfigError("公众号名称不能为空")
    wb = _load_wb(path)
    removed = 0
    try:
        if WECHAT_SHEET not in wb.sheetnames:
            raise SourceConfigError(f"Excel文件中缺少 '{WECHAT_SHEET}' sheet")
        ws = wb[WECHAT_SHEET]
        headers = _header_map(ws, WECHAT_HEADERS[:1])
        for row_idx in range(ws.max_row, 1, -1):
            if _normalize(ws.cell(row_idx, headers["公众号名称"]).value) == name:
                ws.delete_rows(row_idx)
                removed += 1
        if removed:
            _save_wb(wb, path)
        return removed
    finally:
        wb.close()


def _podcast_accounts_from_yaml() -> list[PodcastAccount]:
    accounts = []
    for raw in cfg("podcast.accounts", []) or []:
        if isinstance(raw, str):
            url = _normalize(raw)
            name = ""
            short_name = ""
        else:
            url = _normalize(raw.get("url"))
            name = _normalize(raw.get("name"))
            short_name = _normalize(raw.get("short_name") or name)
        if url:
            accounts.append(PodcastAccount(name=name or url, url=url, short_name=short_name or name or "播客"))
    return accounts


def ensure_podcast_sheet(path: str | Path | None = None) -> bool:
    """Create podcast_account from config.yaml podcast.accounts when absent."""
    file_path = _config_path(path)
    _ensure_openpyxl()
    if file_path.exists():
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
    try:
        if PODCAST_SHEET in wb.sheetnames:
            _header_map(wb[PODCAST_SHEET], PODCAST_HEADERS)
            return False
        ws = wb.create_sheet(PODCAST_SHEET)
        ws.append(list(PODCAST_HEADERS))
        for account in _podcast_accounts_from_yaml():
            ws.append([account.name, account.url, account.short_name])
        _save_wb(wb, path, backup=file_path.exists())
        return True
    finally:
        wb.close()


def load_podcast_accounts(path: str | Path | None = None) -> list[dict[str, str]]:
    ensure_podcast_sheet(path)
    wb = _load_wb(path)
    try:
        ws = wb[PODCAST_SHEET]
        headers = _header_map(ws, PODCAST_HEADERS)
        accounts: list[dict[str, str]] = []
        seen_names: set[str] = set()
        seen_urls: set[str] = set()
        for row_idx in range(2, ws.max_row + 1):
            name = _normalize(ws.cell(row_idx, headers["播客名称"]).value)
            url = _normalize(ws.cell(row_idx, headers["url"]).value)
            short_name = _normalize(ws.cell(row_idx, headers["简称"]).value)
            if not any([name, url, short_name]):
                continue
            if not name or not url or not short_name:
                raise SourceConfigError(f"{PODCAST_SHEET} 第 {row_idx} 行播客名称、url、简称均为必填")
            if name in seen_names:
                raise SourceConfigError(f"{PODCAST_SHEET} 存在重复播客名称: {name}")
            if url in seen_urls:
                raise SourceConfigError(f"{PODCAST_SHEET} 存在重复播客 URL: {url}")
            seen_names.add(name)
            seen_urls.add(url)
            accounts.append({"name": name, "url": url, "short_name": short_name})
        return accounts
    finally:
        wb.close()


def _validate_podcast_url(url: str) -> None:
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or "xiaoyuzhoufm.com" not in parsed.netloc:
        raise SourceConfigError(f"小宇宙播客 URL 无效: {url}")


def add_podcast_account(name: str, url: str, short_name: str, path: str | Path | None = None) -> None:
    name = _normalize(name)
    url = _normalize(url)
    short_name = _normalize(short_name)
    if not name or not url or not short_name:
        raise SourceConfigError("播客名称、url、简称不能为空")
    _validate_podcast_url(url)
    ensure_podcast_sheet(path)
    wb = _load_wb(path)
    try:
        ws = _ensure_sheet(wb, PODCAST_SHEET, PODCAST_HEADERS)
        headers = _header_map(ws, PODCAST_HEADERS)
        for row_idx in range(2, ws.max_row + 1):
            existing_name = _normalize(ws.cell(row_idx, headers["播客名称"]).value)
            existing_url = _normalize(ws.cell(row_idx, headers["url"]).value)
            if existing_name == name:
                raise SourceConfigError(f"播客名称已存在: {name}")
            if existing_url == url:
                raise SourceConfigError(f"播客 URL 已存在: {url}")
        ws.append([name, url, short_name])
        _save_wb(wb, path)
    finally:
        wb.close()


def remove_podcast_account(query: str, path: str | Path | None = None) -> int:
    query = _normalize(query)
    if not query:
        raise SourceConfigError("播客名称或 URL 不能为空")
    ensure_podcast_sheet(path)
    wb = _load_wb(path)
    removed = 0
    try:
        ws = wb[PODCAST_SHEET]
        headers = _header_map(ws, PODCAST_HEADERS[:2])
        for row_idx in range(ws.max_row, 1, -1):
            name = _normalize(ws.cell(row_idx, headers["播客名称"]).value)
            url = _normalize(ws.cell(row_idx, headers["url"]).value)
            if query in {name, url}:
                ws.delete_rows(row_idx)
                removed += 1
        if removed:
            _save_wb(wb, path)
        return removed
    finally:
        wb.close()


def _tag_options_from_ws(ws) -> dict[str, list[str]]:
    headers = _header_map(ws, TAG_HEADERS)
    companies = []
    industries = []
    for row_idx in range(2, ws.max_row + 1):
        companies.append(_normalize(ws.cell(row_idx, headers["公司"]).value))
        industries.append(_normalize(ws.cell(row_idx, headers["行业"]).value))
    return {"公司": _dedupe_keep_order(companies), "行业": _dedupe_keep_order(industries)}


def load_tag_options(path: str | Path | None = None) -> dict[str, list[str]]:
    wb = _load_wb(path)
    try:
        if TAG_SHEET not in wb.sheetnames:
            raise SourceConfigError(f"Excel文件中缺少 '{TAG_SHEET}' sheet")
        return _tag_options_from_ws(wb[TAG_SHEET])
    finally:
        wb.close()


def _rewrite_tag_sheet(ws, companies: list[str], industries: list[str]) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    max_len = max(len(companies), len(industries))
    for idx in range(max_len):
        ws.append([
            companies[idx] if idx < len(companies) else "",
            industries[idx] if idx < len(industries) else "",
        ])


def add_tag_option(kind: str, value: str, path: str | Path | None = None) -> bool:
    kind = _normalize(kind)
    value = _normalize(value)
    if kind not in TAG_HEADERS:
        raise SourceConfigError("标签类型只能是 公司 或 行业")
    if not value:
        return False
    wb = _load_wb(path)
    try:
        ws = _ensure_sheet(wb, TAG_SHEET, TAG_HEADERS)
        options = _tag_options_from_ws(ws)
        if value in options[kind]:
            return False
        options[kind].append(value)
        _rewrite_tag_sheet(ws, options["公司"], options["行业"])
        _save_wb(wb, path)
        return True
    finally:
        wb.close()


def remove_tag_option(kind: str, value: str, path: str | Path | None = None) -> int:
    kind = _normalize(kind)
    value = _normalize(value)
    if kind not in TAG_HEADERS:
        raise SourceConfigError("标签类型只能是 公司 或 行业")
    if not value:
        raise SourceConfigError("标签不能为空")
    wb = _load_wb(path)
    try:
        ws = _ensure_sheet(wb, TAG_SHEET, TAG_HEADERS)
        options = _tag_options_from_ws(ws)
        before = len(options[kind])
        options[kind] = [item for item in options[kind] if item != value]
        removed = before - len(options[kind])
        if removed:
            _rewrite_tag_sheet(ws, options["公司"], options["行业"])
            _save_wb(wb, path)
        return removed
    finally:
        wb.close()
