from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


def _write_yaml_config(path: Path):
    path.write_text(
        """
wechat:
  category_mapping:
    投资: "1-投资"
    商业: "3-商业"
    其他: "6-其他"
podcast:
  accounts:
    - name: "知行小酒馆"
      url: "https://www.xiaoyuzhoufm.com/podcast/6013f9f58e2f7ee375cf4216"
      short_name: "酒馆"
""".strip(),
        encoding="utf-8",
    )


def _patch_config(monkeypatch, tmp_path):
    from investment_system.common.config import config_loader

    config_file = tmp_path / "config.yaml"
    _write_yaml_config(config_file)
    monkeypatch.setattr(config_loader, "_CONFIG_FILE", str(config_file))
    config_loader.reload_config()
    return config_file


def _make_workbook(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "wechat_account"
    ws.append(["公众号名称", "简称", "分类", "单篇/聚合"])
    ws.append(["晚点LatePost", "晚点", "商业", "单篇"])
    tag_ws = wb.create_sheet("memo_tag_options")
    tag_ws.append(["公司", "行业"])
    tag_ws.append(["携程", "宏观"])
    tag_ws.append(["携程", "策略"])
    wb.save(path)
    wb.close()


def test_wechat_accounts_read_add_remove_and_backup(tmp_path, monkeypatch):
    _patch_config(monkeypatch, tmp_path)
    from investment_system.common.config import source_config

    xlsx = tmp_path / "set_config.xlsx"
    _make_workbook(xlsx)

    assert source_config.load_wechat_accounts(xlsx) == [
        {"name": "晚点LatePost", "short_name": "晚点", "category": "商业", "mode": "单篇"}
    ]

    source_config.add_wechat_account("投资聚义厅", "聚义", "投资", path=xlsx)
    names = [item["name"] for item in source_config.load_wechat_accounts(xlsx)]
    assert names == ["晚点LatePost", "投资聚义厅"]
    assert list(tmp_path.glob("set_config.xlsx.*.bak"))

    with pytest.raises(source_config.SourceConfigError, match="公众号已存在"):
        source_config.add_wechat_account("投资聚义厅", "聚义", "投资", path=xlsx)
    with pytest.raises(source_config.SourceConfigError, match="公众号分类无效"):
        source_config.add_wechat_account("坏分类", "坏", "生活", path=xlsx)

    assert source_config.remove_wechat_account("晚点LatePost", xlsx) == 1
    assert [item["name"] for item in source_config.load_wechat_accounts(xlsx)] == ["投资聚义厅"]


def test_podcast_sheet_is_migrated_from_yaml_and_enforced(tmp_path, monkeypatch):
    _patch_config(monkeypatch, tmp_path)
    from investment_system.common.config import source_config

    xlsx = tmp_path / "set_config.xlsx"
    _make_workbook(xlsx)

    accounts = source_config.load_podcast_accounts(xlsx)
    assert accounts == [
        {
            "name": "知行小酒馆",
            "url": "https://www.xiaoyuzhoufm.com/podcast/6013f9f58e2f7ee375cf4216",
            "short_name": "酒馆",
        }
    ]

    wb = load_workbook(xlsx)
    assert "podcast_account" in wb.sheetnames
    wb.close()

    source_config.add_podcast_account(
        "晚点聊",
        "https://www.xiaoyuzhoufm.com/podcast/61933ace1b4320461e91fd55",
        "晚点",
        xlsx,
    )
    assert len(source_config.load_podcast_accounts(xlsx)) == 2

    with pytest.raises(source_config.SourceConfigError, match="播客 URL 已存在"):
        source_config.add_podcast_account(
            "重复URL",
            "https://www.xiaoyuzhoufm.com/podcast/61933ace1b4320461e91fd55",
            "重复",
            xlsx,
        )
    with pytest.raises(source_config.SourceConfigError, match="URL 无效"):
        source_config.add_podcast_account("坏链接", "https://example.com/podcast/1", "坏", xlsx)

    assert source_config.remove_podcast_account("晚点聊", xlsx) == 1
    assert [item["name"] for item in source_config.load_podcast_accounts(xlsx)] == ["知行小酒馆"]


def test_tag_options_add_remove_and_dedupe(tmp_path, monkeypatch):
    _patch_config(monkeypatch, tmp_path)
    from investment_system.common.config import source_config

    xlsx = tmp_path / "set_config.xlsx"
    _make_workbook(xlsx)

    assert source_config.load_tag_options(xlsx) == {"公司": ["携程"], "行业": ["宏观", "策略"]}

    assert source_config.add_tag_option("公司", "美团", xlsx) is True
    assert source_config.add_tag_option("公司", "美团", xlsx) is False
    assert source_config.add_tag_option("行业", "", xlsx) is False
    assert source_config.load_tag_options(xlsx)["公司"] == ["携程", "美团"]

    assert source_config.remove_tag_option("行业", "宏观", xlsx) == 1
    assert source_config.load_tag_options(xlsx)["行业"] == ["策略"]


def test_missing_sheet_or_header_reports_clear_error(tmp_path, monkeypatch):
    _patch_config(monkeypatch, tmp_path)
    from investment_system.common.config import source_config

    xlsx = tmp_path / "broken.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "wechat_account"
    ws.append(["公众号名称", "简称"])
    wb.save(xlsx)
    wb.close()

    with pytest.raises(source_config.SourceConfigError, match="缺少必要表头"):
        source_config.load_wechat_accounts(xlsx)
    with pytest.raises(source_config.SourceConfigError, match="缺少 'memo_tag_options'"):
        source_config.load_tag_options(xlsx)
